from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import List, Optional
from datetime import datetime, timedelta
import io
import csv
import re
import pytz

from app.db.session import get_db
from app.db.models import Sale, Expense
from app.schemas.sale import SaleCreate, SaleUpdate, SaleResponse

router = APIRouter()

# NZ Timezone
NZ_TZ = pytz.timezone('Pacific/Auckland')

# Size and syrup modifiers that are not counted as separate items
SIZE_MODIFIERS = {
    'grande', 'venti', 'tall', 'short', 'medium', 'large', 'small',
    'syrup', 'ube syrup', 'caramel syrup', 'vanilla syrup', 'hazelnut syrup',
    'chocolate syrup', 'cinnamon syrup', 'honey', 'whipped cream'
}

def parse_item_and_quantity(item_name: str):
    """
    Parse item name to extract quantity and clean name.
    Handles formats like:
    - "1x Item"
    - "Item ×3"
    - "1x Item1 2x Item2"
    - "Ube CB 2x Twilight Cloud" -> [(Ube CB, 1), (Twilight Cloud, 2)]
    - "Item1 + Item2"
    Returns list of (clean_name, quantity) tuples
    """
    if not item_name or not item_name.strip():
        return []
    
    results = []
    
    # Split by commas first (for "Item1 ×3, Item2 ×3")
    comma_items = [i.strip() for i in item_name.split(',')]
    
    for comma_item in comma_items:
        if not comma_item:
            continue
        
        # Skip pure size/modifier items
        if comma_item.lower() in SIZE_MODIFIERS:
            continue
        
        # Skip if item is just a syrup or ends with "syrup"
        if 'syrup' in comma_item.lower() or comma_item.lower() in SIZE_MODIFIERS:
            continue
        
        # Split by " + " or " and "
        sub_items = re.split(r'\s*\+\s*|\s+and\s+', comma_item)
        
        for sub_item in sub_items:
            sub_item = sub_item.strip()
            if not sub_item:
                continue
            
            # Skip if it's a syrup or modifier
            if 'syrup' in sub_item.lower() or sub_item.lower() in SIZE_MODIFIERS:
                continue
            
            # Try "1x Item" format at beginning
            match_beginning = re.match(r'^(\d+)\s*[×x]\s+(.+)$', sub_item)
            if match_beginning:
                qty = int(match_beginning.group(1))
                name = match_beginning.group(2).strip()
                # Remove trailing size modifiers
                for size in SIZE_MODIFIERS:
                    name = re.sub(r'\s+' + size + r'\s*$', '', name, flags=re.IGNORECASE)
                name = name.strip()
                if name and name.lower() not in SIZE_MODIFIERS and 'syrup' not in name.lower():
                    results.append((name, qty))
                continue
            
            # Try "Item ×3" format at end
            match_end = re.search(r'^(.+?)\s+[×x]\s?(\d+)$', sub_item)
            if match_end:
                name = match_end.group(1).strip()
                qty = int(match_end.group(2))
                # Remove trailing size modifiers
                for size in SIZE_MODIFIERS:
                    name = re.sub(r'\s+' + size + r'\s*$', '', name, flags=re.IGNORECASE)
                name = name.strip()
                if name and name.lower() not in SIZE_MODIFIERS and 'syrup' not in name.lower():
                    results.append((name, qty))
                continue
            
            # No quantity notation - default to 1
            sub_item_clean = sub_item
            for size in SIZE_MODIFIERS:
                sub_item_clean = re.sub(r'\s+' + size + r'\s*$', '', sub_item_clean, flags=re.IGNORECASE)
            sub_item_clean = sub_item_clean.strip()
            if sub_item_clean and sub_item_clean.lower() not in SIZE_MODIFIERS and 'syrup' not in sub_item_clean.lower():
                results.append((sub_item_clean, 1))
    
    return results


@router.post("/sale", response_model=SaleResponse, status_code=201)
async def create_sale(payload: SaleCreate, db: AsyncSession = Depends(get_db)):
    """Record a completed sale."""
    # Use provided date (manual/past entry) or current NZ time
    if payload.date:
        sale_date = payload.date
    else:
        # Get current time in NZ timezone
        sale_date = datetime.now(NZ_TZ).replace(tzinfo=None)

    # Calculate order number for the sale's specific day
    day_start = sale_date.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)
    count_result = await db.execute(
        select(func.count()).select_from(Sale).where(
            Sale.date >= day_start, Sale.date < day_end
        )
    )
    daily_order_number = (count_result.scalar() or 0) + 1

    # Build items list — use description as single item name for manual entries
    items = [item.model_dump() for item in payload.items]
    if not items and payload.description:
        items = [{"id": "manual", "name": payload.description, "price": round(payload.subtotal, 2), "quantity": 1}]

    sale = Sale(
        date=sale_date,
        items=items,
        subtotal=round(payload.subtotal, 2),
        discount=round(payload.discount or 0.0, 2),
        payment_method=payload.payment_method,
        daily_order_number=daily_order_number,
        customer_name=payload.customer_name.strip() if payload.customer_name else None,
    )
    db.add(sale)
    await db.commit()
    await db.refresh(sale)
    return sale


@router.get("/sales", response_model=List[SaleResponse])
async def get_sales(
    date: Optional[str] = Query(None, description="Filter by date YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db)
):
    """Return sales, newest first. Optionally filter by date (YYYY-MM-DD)."""
    query = select(Sale).order_by(Sale.date.desc())
    if date:
        try:
            d = datetime.strptime(date, "%Y-%m-%d")
            query = query.where(Sale.date >= d).where(Sale.date < d + timedelta(days=1))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/sales/export")
async def export_sales(
    date: Optional[str] = Query(None, description="Filter by date YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db)
):
    """Export sales + expenses to an Excel file (.xlsx). Optionally filter by date."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    # Date filter
    date_filter_start = None
    date_filter_end = None
    if date:
        try:
            date_filter_start = datetime.strptime(date, "%Y-%m-%d")
            date_filter_end = date_filter_start + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")

    # Fetch sales
    sales_query = select(Sale).order_by(Sale.date.asc())
    if date_filter_start:
        sales_query = sales_query.where(Sale.date >= date_filter_start).where(Sale.date < date_filter_end)
    sales_result = await db.execute(sales_query)
    sales = sales_result.scalars().all()

    # Fetch expenses
    exp_query = select(Expense).order_by(Expense.date.asc())
    if date_filter_start:
        exp_query = exp_query.where(Expense.date >= date_filter_start).where(Expense.date < date_filter_end)
    exp_result = await db.execute(exp_query)
    expenses = exp_result.scalars().all()

    wb = openpyxl.Workbook()

    # ── Sales sheet ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Sales"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="6B4C3B")  # cafe-brown
    sales_headers = ["Order #", "Name", "Date", "Items", "Quantity", "Subtotal (NZD)", "Discount (NZD)", "Net Total (NZD)", "Payment Method"]
    ws.append(sales_headers)
    for col_idx, _ in enumerate(sales_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for sale in sales:
        items_list = sale.items or []
        item_names = []
        total_quantity = 0
        
        for item in items_list:
            item_name = item.get('name', '')
            item_qty = item.get('quantity', 1)
            
            # Check if quantity is embedded in name (old format) or separate (new format)
            if item_qty == 1 and item_name and any(marker in item_name for marker in ['x ', '×', ' and ', ' + ']):
                # Old format: parse quantity from name
                parsed_items = parse_item_and_quantity(item_name)
                for clean_name, qty in parsed_items:
                    item_names.append(clean_name)
                    total_quantity += qty
            else:
                # New format: use stored quantity directly
                # Skip if it's a size modifier
                if item_name and item_name.lower().strip() in SIZE_MODIFIERS:
                    continue
                
                if item_name and item_name.strip():
                    item_names.append(item_name)
                    total_quantity += item_qty
        
        items_str = ", ".join(item_names)
        
        order_label = f"#{sale.daily_order_number}" if sale.daily_order_number else f"ID {sale.id}"
        discount = round(sale.discount or 0.0, 2)
        net_total = round(sale.subtotal - discount, 2)
        ws.append([
            order_label,
            sale.customer_name or "",
            sale.date.strftime("%Y-%m-%d %H:%M:%S") if sale.date else "",
            items_str,
            total_quantity if total_quantity > 0 else "",
            round(sale.subtotal, 2),
            discount if discount else "",
            net_total,
            sale.payment_method,
        ])

    # Totals row
    # Calculate totals using stored/parsed quantities, skip size modifiers
    total_net = sum(round((s.subtotal or 0) - (s.discount or 0), 2) for s in sales)
    total_qty = 0
    for sale in sales:
        for item in (sale.items or []):
            item_name = item.get('name', '')
            item_qty = item.get('quantity', 1)
            
            # Check if quantity is embedded in name (old format) or separate (new format)
            if item_qty == 1 and item_name and any(marker in item_name for marker in ['x ', '×', ' and ', ' + ']):
                # Old format: parse quantity from name
                parsed_items = parse_item_and_quantity(item_name)
                for _, qty in parsed_items:
                    total_qty += qty
            else:
                # New format: use stored quantity directly
                if item_name and item_name.lower().strip() not in SIZE_MODIFIERS:
                    total_qty += item_qty
    
    ws.append(["", "", "", "TOTAL", total_qty if total_qty > 0 else "", "", round(total_net, 2), ""])
    total_row = ws.max_row
    for col_idx in range(1, len(sales_headers) + 1):
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

    # Column widths
    for col_idx, width in enumerate([10, 18, 20, 40, 10, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Expenses sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Expenses")

    exp_fill = PatternFill(fill_type="solid", fgColor="C0392B")  # red
    exp_headers = ["Date", "Category", "Description", "Paid From", "Amount (NZD)"]
    ws2.append(exp_headers)
    for col_idx, _ in enumerate(exp_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = exp_fill
        cell.alignment = Alignment(horizontal="center")

    for exp in expenses:
        paid_label = "Own Money" if getattr(exp, "paid_from", None) == "own" else "Sales Money"
        ws2.append([
            exp.date.strftime("%Y-%m-%d %H:%M:%S") if exp.date else "",
            exp.category,
            exp.description or "",
            paid_label,
            round(exp.amount, 2),
        ])

    # Totals row
    total_exp = round(sum(e.amount for e in expenses), 2)
    ws2.append(["TOTAL", "", "", "", total_exp])
    total_row2 = ws2.max_row
    for col_idx in range(1, len(exp_headers) + 1):
        ws2.cell(row=total_row2, column=col_idx).font = Font(bold=True)

    for col_idx, width in enumerate([20, 14, 40, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    label = f"_{date}" if date else ""
    filename = f"stllhaus_export{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.put("/sale/{sale_id}", response_model=SaleResponse)
async def update_sale(
    sale_id: int,
    payload: SaleUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Update an existing sale (date, amount, payment method, customer, description)."""
    result = await db.execute(select(Sale).where(Sale.id == sale_id))
    sale = result.scalar_one_or_none()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    sale_date = payload.date  # already local time, no UTC conversion needed
    sale.date = sale_date
    sale.subtotal = round(payload.subtotal, 2)
    sale.discount = round(payload.discount or 0.0, 2)
    sale.payment_method = payload.payment_method
    sale.customer_name = payload.customer_name.strip() if payload.customer_name else None

    # Update description in items if it's a manual entry or if description provided
    if payload.description:
        sale.items = [{"id": "manual", "name": payload.description, "price": round(payload.subtotal, 2), "quantity": 1}]
    elif sale.items and payload.description == "":
        sale.items = [{"id": "manual", "name": "", "price": round(payload.subtotal, 2), "quantity": 1}]

    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(sale, "items")
    await db.commit()
    await db.refresh(sale)
    return sale


@router.delete("/sale/{sale_id}", status_code=200)
async def delete_sale(sale_id: int, db: AsyncSession = Depends(get_db)):
    """Delete a single sale by ID."""
    result = await db.execute(select(Sale).where(Sale.id == sale_id))
    sale = result.scalar_one_or_none()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    await db.delete(sale)
    await db.commit()
    return {"ok": True}


@router.delete("/sales", status_code=200)
async def reset_sales(db: AsyncSession = Depends(get_db)):
    """Delete ALL sales records — use for trial/testing resets only."""
    from sqlalchemy import delete
    await db.execute(delete(Sale))
    await db.commit()
    return {"message": "All sales deleted"}


@router.get("/sales/export/csv")
async def export_sales_csv(
    date: Optional[str] = Query(None, description="Filter by date YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db)
):
    """Export sales to CSV format: Order #, Name, Date, Items, Quantity, Subtotal, Discount, Net Total, Payment Method"""
    # Date filter
    date_filter_start = None
    date_filter_end = None
    if date:
        try:
            date_filter_start = datetime.strptime(date, "%Y-%m-%d")
            date_filter_end = date_filter_start + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")

    # Fetch sales
    sales_query = select(Sale).order_by(Sale.date.asc())
    if date_filter_start:
        sales_query = sales_query.where(Sale.date >= date_filter_start).where(Sale.date < date_filter_end)
    sales_result = await db.execute(sales_query)
    sales = sales_result.scalars().all()

    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(["Order #", "Name", "Date", "Items", "Quantity", "Subtotal (NZD)", "Discount (NZD)", "Net Total (NZD)", "Payment Method"])
    
    # Write sales data - one row per sale
    for sale in sales:
        order_label = f"#{sale.daily_order_number}" if sale.daily_order_number else f"ID {sale.id}"
        sale_date = sale.date.strftime("%Y-%m-%d %H:%M:%S") if sale.date else ""
        items = sale.items or []
        
        # Parse items and calculate total quantity
        item_names = []
        total_quantity = 0
        
        for item in items:
            item_name = item.get("name", "")
            item_qty = item.get("quantity", 1)
            
            # Check if quantity is embedded in name (old format) or separate (new format)
            if item_qty == 1 and item_name and any(marker in item_name for marker in ['x ', '×', ' and ', ' + ']):
                # Old format: parse quantity from name
                parsed_items = parse_item_and_quantity(item_name)
                for clean_name, qty in parsed_items:
                    item_names.append(clean_name)
                    total_quantity += qty
            else:
                # New format: use stored quantity directly
                if item_name and item_name.lower().strip() in SIZE_MODIFIERS:
                    continue
                
                if item_name and item_name.strip():
                    item_names.append(item_name)
                    total_quantity += item_qty
        
        items_str = ", ".join(item_names)
        discount = round(sale.discount or 0.0, 2)
        net_total = round((sale.subtotal or 0) - discount, 2)
        
        writer.writerow([
            order_label,
            sale.customer_name or "",
            sale_date,
            items_str,
            total_quantity if total_quantity > 0 else "",
            f"{sale.subtotal or 0:.2f}",
            f"{discount:.2f}" if discount else "",
            f"{net_total:.2f}",
            sale.payment_method,
        ])
    
    # Write totals row
    total_subtotal = sum(s.subtotal or 0 for s in sales)
    total_discount = sum(s.discount or 0 for s in sales)
    total_net = sum(round((s.subtotal or 0) - (s.discount or 0), 2) for s in sales)
    
    # Calculate total quantity
    total_qty = 0
    for sale in sales:
        for item in (sale.items or []):
            item_name = item.get('name', '')
            item_qty = item.get('quantity', 1)
            
            # Check if quantity is embedded in name (old format) or separate (new format)
            if item_qty == 1 and item_name and any(marker in item_name for marker in ['x ', '×', ' and ', ' + ']):
                # Old format: parse quantity from name
                parsed_items = parse_item_and_quantity(item_name)
                for _, qty in parsed_items:
                    total_qty += qty
            else:
                # New format: use stored quantity directly
                if item_name and item_name.lower().strip() not in SIZE_MODIFIERS:
                    total_qty += item_qty
    
    writer.writerow(["", "", "", "TOTAL", total_qty if total_qty > 0 else "", f"{total_subtotal:.2f}", f"{total_discount:.2f}", f"{total_net:.2f}", ""])

    stream = io.BytesIO(output.getvalue().encode('utf-8'))
    label = f"_{date}" if date else ""
    filename = f"stllhaus_sales{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/sales/import/csv", status_code=201)
async def import_sales_csv(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Import sales from CSV format: Order #, Name, Date, Items, Quantity, Subtotal, Discount, Net Total, Payment Method"""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV file")
    
    try:
        contents = await file.read()
        text = contents.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="CSV file is empty")
        
        # Expected columns
        expected_cols = {"Order #", "Name", "Date", "Items", "Quantity", "Subtotal (NZD)", "Discount (NZD)", "Net Total (NZD)", "Payment Method"}
        provided_cols = set(reader.fieldnames)
        
        if not expected_cols.issubset(provided_cols):
            missing = expected_cols - provided_cols
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing)}"
            )
        
        imported_count = 0
        
        for row_idx, row in enumerate(reader, 1):
            # Skip empty rows or total rows
            if not row.get("Date") or row.get("Items", "").strip().upper() == "TOTAL":
                continue
            
            try:
                # Parse date
                date_str = row.get("Date", "").strip()
                if not date_str:
                    continue
                sale_date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                
                # Parse fields
                items_str = row.get("Items", "").strip()
                quantity = int(row.get("Quantity", 1))
                subtotal = float(row.get("Subtotal (NZD)", 0))
                discount = float(row.get("Discount (NZD)", 0))
                payment_method = row.get("Payment Method", "Cash").strip() or "Cash"
                customer_name = row.get("Name", "").strip()
                
                if not items_str:
                    continue
                
                # Split items by comma and create item objects
                item_names = [item.strip() for item in items_str.split(",")]
                items = []
                price_per_item = subtotal / quantity if quantity > 0 else 0
                
                for item_name in item_names:
                    if item_name:
                        items.append({
                            "id": item_name.lower().replace(" ", "_"),
                            "name": item_name,
                            "quantity": 1,
                            "price": price_per_item
                        })
                
                # Calculate order number for the day
                day_start = sale_date.replace(hour=0, minute=0, second=0, microsecond=0)
                day_end = day_start + timedelta(days=1)
                count_result = await db.execute(
                    select(func.count()).select_from(Sale).where(
                        Sale.date >= day_start, Sale.date < day_end
                    )
                )
                daily_order_number = (count_result.scalar() or 0) + 1
                
                sale = Sale(
                    date=sale_date,
                    items=items,
                    subtotal=round(subtotal, 2),
                    discount=round(discount, 2),
                    payment_method=payment_method,
                    daily_order_number=daily_order_number,
                    customer_name=customer_name or None,
                )
                db.add(sale)
                imported_count += 1
                
            except ValueError as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid data in row {row_idx}: {str(e)}"
                )
        
        await db.commit()
        return {
            "message": f"Successfully imported {imported_count} sales",
            "count": imported_count
        }
        
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="File must be UTF-8 encoded")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing CSV: {str(e)}")
